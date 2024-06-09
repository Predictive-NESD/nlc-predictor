# Import Libraries
import argparse
import os
from datetime import datetime
import pandas as pd
import numpy as np
import shap
from IPython.display import display, HTML
import matplotlib
import matplotlib.pyplot as plt
import os
import io
from openpyxl import load_workbook, Workbook
import tensorflow as tf
from tensorflow import keras
from sklearn import preprocessing
from sklearn.model_selection import train_test_split

# Functions
def parse_arguments():
    parser = argparse.ArgumentParser(description='Run the prediction model on input Excel file and save the output.')
    parser.add_argument('input_file', type=str, help='Input Excel file name')
    parser.add_argument('output_file', type=str, help='Output Excel file name')
    return parser.parse_args()

# read file & load model
def read_excel_file(file_name):
    file_path = '../input/{}.xlsx'.format(file_name)
    return pd.read_excel(file_path)

def read_csv_file(file_name):
   file_path = '../input/{}.csv'.format(file_name)
   return pd.read_csv(file_path)

# def read_case(file_name, case_name):
#   case_df = pd.read_excel(
#     io='../input/{}.xlsx'.format(file_name),
#     sheet_name=case_name,
#     usecols='A:AQ',
#     nrows=28
#   )
#   case_df['Country Name'] = case_df['Country Name'].fillna('Thailand')
#   case_df.rename(columns={'Changes in inventory index': 'Changes in inventories (current US$)'}, inplace=True)
#   return case_df
def read_case(file_name):
    case_df = pd.read_excel(
       io=file_name,
       sheet_name=0,  # Read the first sheet/tab
       usecols='A:AQ',
       nrows=28
    )
    case_df['Country Name'] = case_df['Country Name'].fillna('Thailand')
    case_df.rename(columns={'Changes in inventory index': 'Changes in inventories (current US$)'}, inplace=True)
    return case_df

def load_model(model_name):
  return keras.models.load_model(
    '../model/{}.keras'.format(model_name))

# calculate
def calculate_index(df, column_names):
  for column in column_names:
    print(f'Creating index for {column}')
    base_year = df['Year'].min()
    base_year_value = df[df['Year'] == base_year][['Country Name', column]].set_index('Country Name')
    df = df.merge(base_year_value, left_on='Country Name', right_index=True, suffixes=('', ' base'))
    index_column_name = f'{column} Index'
    df[index_column_name] = (df[column] / df[f'{column} base']) * 100
    df = df.drop([f'{column} base'], axis=1)
  return df

def calculate_transformed(df, value, column_names):
  for column in column_names:
    print(f'Creating transformed for {column}')
    transformed_column_name = f'{column} Transformed'
    df[transformed_column_name] = value / df[column]
  return df

# add column name to calculate index & transformed here
def calculate_all(case_df):
  new_case_df = calculate_all_index(case_df)
  new_case_df = calculate_all_transformed(new_case_df)
  return new_case_df

def calculate_all_index(case_df):
  new_case_df = calculate_index(case_df, ['Urban Population', 'Lending interest rate (%)',
                              'Railways, goods transported (million ton-km)',
                              'Container port traffic (TEU: 20 foot equivalent units)',
                              'Official exchange rate (LCU per US$, period average)',
                              'Foreign direct investment, net inflows (BoP, current US$)',
                              'GDP (constant 2015 US$)',
                              'Pump price for diesel fuel (US$ per liter)',
                              'Road freight transport',
                              'Roads total network (km)',
                              'Rail lines (total route-km)',
                              'Air transport, freight (million ton-km)'])
  return new_case_df

def calculate_all_transformed(case_df):
  new_case_df = calculate_transformed(case_df, 5, ['LPI Infra',
                                                      'LPI International Shipments',
                                                      'LPI Logistics Competence'])
  new_case_df = calculate_transformed(new_case_df, 7, ['Quality of roads, 1(low) - 7(high)',
                                                      'Quality of railroad infrastructure, 1(low) - 7(high)',
                                                      'Quality of air transport infrastructure, 1(low) - 7(high)'])
  new_case_df = calculate_transformed(new_case_df, 10000, ['Urban Population Index',
                                                          'Import volume index (2000 = 100)',
                                                          'Export volume index (2000 = 100)',
                                                          'Changes in inventories (current US$)',
                                                          'Railways, goods transported (million ton-km) Index',
                                                          'Container port traffic (TEU: 20 foot equivalent units) Index',
                                                          'Road freight transport Index',
                                                          'Roads total network (km) Index',
                                                          'Rail lines (total route-km) Index',
                                                          'Air transport, freight (million ton-km) Index'])
  return new_case_df

# predict & save
def predict_and_display(best_model, normalized_features, index_name, output_file, case_name):
    ann_predictions = best_model.predict(normalized_features)
    print(f"==========Prediction {index_name} Index ==========")
    for pred in ann_predictions:
        print("{}".format(pred[0]))

    excel_file_path = output_file

    # Check if the excel file already exists
    if os.path.exists(excel_file_path):
        wb = load_workbook(excel_file_path)
        if 'Sheet' in wb.sheetnames:
          del wb['Sheet']
    else:
        # If the Excel file doesn't exist, create a new workbook
        wb = Workbook()

    ws = wb[case_name] if case_name in wb.sheetnames else wb.create_sheet(case_name)

    # Add the Year column if it's a new sheet
    if not ws['A1'].value:
        ws['A1'] = 'Year'
        for i in range(len(ann_predictions)):
            ws.cell(row=i + 2, column=1, value=2007 + i)

    if all(ws.cell(row=1, column=i).value != f'{index_name} Index' for i in range(1, ws.max_column + 1)):
        ws.cell(row=1, column=ws.max_column + 1, value=f'{index_name} Index')

    index_column_index = next((i for i in range(1, ws.max_column + 1) if ws.cell(row=1, column=i).value == f'{index_name} Index'), None)

    for i, pred in enumerate(ann_predictions):
        ws.cell(row=i + 2, column=index_column_index, value=pred[0])

    wb.save(excel_file_path)

# create new folder
def create_new_folder(base_path, folder_name):
    new_folder_path = os.path.join(base_path, folder_name)
    os.makedirs(new_folder_path, exist_ok=True)
    return new_folder_path

### save SHAP plots and SHAP values
def save_force_plot(force_plot, year, index_name, df_name, folder_path):
    shap.initjs()
    file_path = os.path.join(folder_path, f"force_plot_{year}_{index_name}_{df_name}.html")
    shap.save_html(file_path, force_plot)
    print(f"Force plot saved for year {year} {index_name} Index at: {file_path}")

def calculate_shap_and_display(best_model, normalized_features, feature_names, df, index_name, df_name, base_path):
    # Create a new parent folder for SHAP results
    parent_folder_name = f"shap_{df_name}"
    parent_folder_path = create_new_folder(base_path, parent_folder_name)

    # Create subfolders for SHAP plots and SHAP values
    shap_plots_folder = create_new_folder(parent_folder_path, "shap_plots")
    shap_values_folder = create_new_folder(parent_folder_path, "shap_values")

    explainer = shap.KernelExplainer(best_model, normalized_features, feature_names=feature_names)
    shap_values = explainer.shap_values(normalized_features)
    explainer_as_shap_values = explainer(normalized_features)

    shap.initjs()
    skip = 0
    base_value = explainer.expected_value

    shap_dfs = {}

    for rec_ind in range(len(normalized_features)):
        if rec_ind < skip:
            continue

        year = df.loc[rec_ind]['Year']
        if 2021 <= year <= 2033:
            print("Year:", year)
            force_plot = shap.force_plot(base_value, shap_values[0][rec_ind], normalized_features[rec_ind], feature_names=feature_names, show=False)
            display(force_plot)
            save_force_plot(force_plot, year, index_name, df_name, shap_plots_folder)

            if year not in shap_dfs:
                shap_dfs[year] = pd.DataFrame(columns=['Feature'] + [f'{year}'])

            shap_values_for_record = shap_values[0][rec_ind]

            for i, feature_name in enumerate(feature_names):
                shap_value_for_feature = shap_values_for_record[i]
                if feature_name not in shap_dfs[year]['Feature'].values:
                    new_row = pd.DataFrame({'Feature': [feature_name], f'{year}': [shap_value_for_feature]})
                    shap_dfs[year] = pd.concat([shap_dfs[year], new_row], ignore_index=True)
                else:
                    shap_dfs[year].loc[shap_dfs[year]['Feature'] == feature_name, f'{year}'] = shap_value_for_feature

    for year, shap_df in shap_dfs.items():
      print(f"Year: {year}")
      pd.options.display.float_format = '{:.10f}'.format
      print(shap_df)
      
    excel_file_path = os.path.join(shap_values_folder, f'shap_values_{index_name}_{df_name}.xlsx')

    if os.path.exists(excel_file_path):
        wb = load_workbook(excel_file_path)
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
    else:
        wb = Workbook()

    ws = wb[df_name] if df_name in wb.sheetnames else wb.create_sheet(df_name)

    # Write header row with index name and years
    ws.append([f'Feature - {index_name} Index'] + [year for year in shap_dfs.keys()])

    # Write feature names and SHAP values for each year
    for feature_name in feature_names:
        row_data = [feature_name]
        for year in shap_dfs.keys():
            shap_value = shap_dfs[year].loc[shap_dfs[year]['Feature'] == feature_name, f'{year}'].values[0]
            row_data.append(shap_value)
        ws.append(row_data)

    wb.save(excel_file_path)
    print(f"SHAP values saved to {excel_file_path}")


def main():
  args = parse_arguments()
  # Cases & Models Configuration
  # Cases
  case_base = "Test Lag Data 2024"

  # shap path
  shap_path = "../output/shap"

  # input_abc.xlxs
  input_file = args.input_file
  input_file_name = os.path.splitext(args.input_file)[0]
  # case_excel = 'input'
  # case_1 = ''

  # output_xyz.xlxs
  output_file = f'../output/{args.output_file}'

  # Models
  model_name_tc = "tc_index_kt_b2.2_2024_02_02_20_08_42_190891"
  best_model_tc = load_model(model_name_tc)
  model_name_wic = "wic_index_kt_b2.4_2024_02_02_20_08_42_190891"
  best_model_wic = load_model(model_name_wic)
  model_name_ac = "ac_index_kt_b2.4_2024_02_18_05_42_21_089087"
  best_model_ac = load_model(model_name_ac)

  keras.utils.set_random_seed(812)
  tf.config.experimental.enable_op_determinism()

  # Data Scaler
  # data_scaler = preprocessing.StandardScaler()
  data_scaler =  preprocessing.MinMaxScaler()
  # data_scaler = preprocessing.RobustScaler()

  # Transportation cost index
  target_tc = 'Transportation cost index'
  features_tc = [
      'Import volume index (2000 = 100) Transformed',
      'Export volume index (2000 = 100) Transformed',
      'Railways, goods transported (million ton-km) Index Transformed',
      'Container port traffic (TEU: 20 foot equivalent units) Index Transformed',
      'LPI Infra Transformed',
      'LPI International Shipments Transformed',
      'Foreign direct investment, net inflows (BoP, current US$) Index',
      'LPI Logistics Competence Transformed',
      'GDP (constant 2015 US$) Index',
      'Pump price for diesel fuel (US$ per liter) Index',
      'Consumer Price Index',
      'Road freight transport Index Transformed',
      'Roads total network (km) Index Transformed',
      'Rail lines (total route-km) Index Transformed',
      'Air transport, freight (million ton-km) Index Transformed',
      'Quality of roads, 1(low) - 7(high) Transformed',
      'Quality of railroad infrastructure, 1(low) - 7(high) Transformed',
      'Quality of port infrastructure, 1(low) - 7(high)',
      'Quality of air transport infrastructure, 1(low) - 7(high) Transformed'
  ]

  # Warehouse & Inventory cost index
  target_wic = 'WH & Inv cost index'
  features_wic = [
      'Import volume index (2000 = 100) Transformed',
      'Export volume index (2000 = 100) Transformed',
      'Changes in inventories (current US$) Transformed',
      'Lending interest rate (%) Index',
      'LPI Infra Transformed',
      'Foreign direct investment, net inflows (BoP, current US$) Index',
      'LPI Logistics Competence Transformed',
      'GDP (constant 2015 US$) Index',
      'Consumer Price Index',
  ]

  # Admin cost index
  target_ac = 'Admin cost index'
  features_ac = [
      'Import volume index (2000 = 100)',
      'Export volume index (2000 = 100)',
      'Official exchange rate (LCU per US$, period average) Index',
      'Corruption Perceptions Index, 100 = no corruption',
      'LPI Customs',
      'LPI Tracking and Tracing',
      'Foreign direct investment, net inflows (BoP, current US$) Index',
      'LPI Logistics Competence',
      'LPI Timeliness',
      'Consumer Price Index',
  ]

  # Test Data
  df_base = read_csv_file(case_base)
  df_base = calculate_all(df_base)

  # Input Data
  df_case_input = read_case(input_file)
  df_case_input = calculate_all(df_case_input)

  # TC
  df_base[features_tc]
  scaler_tc = data_scaler.fit(df_base[features_tc])
  # print(scaler_tc)

  # Case Input + TC Index
  df_case_input[features_tc]
  normalized_df_case_input_tc = scaler_tc.transform(df_case_input[features_tc])
  normalized_df_case_input_tc.shape
  predict_and_display(best_model_tc, normalized_df_case_input_tc, 'TC', output_file, input_file_name)
  calculate_shap_and_display(best_model_tc, normalized_df_case_input_tc, features_tc, df_case_input, 'TC', input_file_name, shap_path)

  # WIC
  df_base[features_wic]
  scaler_wic = data_scaler.fit(df_base[features_wic])
  # print(scaler_wic)

  # Case Input + WIC Index
  normalized_df_case_input_wic = scaler_wic.transform(df_case_input[features_wic])
  normalized_df_case_input_wic.shape
  predict_and_display(best_model_wic, normalized_df_case_input_wic, 'WIC', output_file, input_file_name)
  calculate_shap_and_display(best_model_wic, normalized_df_case_input_wic, features_wic, df_case_input, 'WIC', input_file_name, shap_path)


  # AC
  df_base[features_ac]
  scaler_ac = data_scaler.fit(df_base[features_ac])
  # print(scaler_ac)

  # Case Input + AC Index
  normalized_df_case_input_ac = scaler_ac.transform(df_case_input[features_ac])
  normalized_df_case_input_ac.shape
  predict_and_display(best_model_ac, normalized_df_case_input_ac, 'AC', output_file, input_file_name)
  calculate_shap_and_display(best_model_ac, normalized_df_case_input_ac, features_ac, df_case_input, 'AC', input_file_name, shap_path)


if __name__ == "__main__":
    main()