# Import Libraries
import argparse
import os
from openpyxl import load_workbook, Workbook
import pandas as pd
from pandas import ExcelFile

# Functions
def parse_arguments():
    parser = argparse.ArgumentParser(description='Merge the output Excel file into a main Excel file as a new sheet.')
    parser.add_argument('output_file', type=str, help='Output Excel file name')
    parser.add_argument('merge_file', type=str, help='Merge Output Excel file name')
    parser.add_argument('sheet_name', type=str, nargs='?', default=None, help='Optional sheet name for the merged data')
    return parser.parse_args()

# read output file
def read_excel_file(file_name):
    file_path = 'output/{}.xlsx'.format(file_name)
    return pd.ExcelFile(file_path)

# merge output file with merge file
def merge_output(output_file_name,merge_file_name, sheet_name=None):
    output_folder_path = 'output'
    output_file = '{}.xlsx'.format(output_file_name)
    merge_folder_path = 'merge'
    merge_file = '{}.xlsx'.format(merge_file_name)
    os.makedirs(output_folder_path, exist_ok=True)
    os.makedirs(merge_folder_path, exist_ok=True)

    merge_path = os.path.join(merge_folder_path, merge_file)

    # Retrieve the ExcelFile object and then parse the first sheet into a DataFrame
    output_excel_file = read_excel_file(output_file_name)
    output_df = pd.read_excel(output_excel_file, sheet_name=0)

    # check if merge excel file is already exist yet
    if os.path.exists(merge_path):
        wb = load_workbook(merge_path)
        if 'Sheet' in wb.sheetnames:
          del wb['Sheet']
    else:
        # if the excel file doesn't exist, create a new workbook
        wb = Workbook()

    # Determine sheet name
    if not sheet_name:
        base_sheet_name = os.path.splitext(output_file_name)[0]
    else:
        base_sheet_name = sheet_name

    sheet_name = base_sheet_name
    sheet_index = 1

    # find a unique sheet name
    while sheet_name in wb.sheetnames:
        sheet_index += 1
        sheet_name = f"{base_sheet_name} ({sheet_index})"

    # remove the default 'Sheet' if it exists and has no data
    if 'Sheet' in wb.sheetnames:
        sheet = wb['Sheet']
        if not any(sheet.iter_rows(min_row=2, max_row=2, min_col=1, max_col=1)):
            del wb['Sheet']

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    for r in dataframe_to_rows(output_df, index=False, header=True):
        ws.append(r)

    wb.save(merge_path)
    print(f"Data from {output_file} merged into {merge_file} as sheet '{sheet_name}'")

def dataframe_to_rows(df, index=True, header=True):
    import itertools
    rows = itertools.chain(
        ([df.columns.values.tolist()] if header else []),
        df.itertuples(index=index, name=None)
    )
    return rows

def main():
    args = parse_arguments()
    output_file_name = os.path.splitext(args.output_file)[0]
    merge_file_name = os.path.splitext(args.merge_file)[0]
    sheet_name = args.sheet_name

    merge_output(output_file_name,merge_file_name, sheet_name)

if __name__ == "__main__":
   main()